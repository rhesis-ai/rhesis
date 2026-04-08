"""File parsers for import.

Reads uploaded files and returns raw row dictionaries.
Delegates to the SDK's normalization logic where possible.
"""

import csv
import io
import json
import logging
from typing import Any, Dict, List, Tuple

logger = logging.getLogger(__name__)


# Maximum sample rows returned during the analyze step
MAX_SAMPLE_ROWS = 5


def detect_format(filename: str) -> str:
    """Detect file format from the filename extension.

    Returns one of: json, jsonl, csv, xlsx
    Raises ValueError for unsupported formats.
    """
    lower = filename.lower()
    if lower.endswith(".jsonl"):
        return "jsonl"
    if lower.endswith(".json"):
        return "json"
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return "xlsx"
    raise ValueError(
        f"Unsupported file format: {filename}. Supported formats: .json, .jsonl, .csv, .xlsx, .xls"
    )


def parse_file(
    file_bytes: bytes,
    file_format: str,
) -> List[Dict[str, Any]]:
    """Parse file bytes into a list of raw row dictionaries.

    Args:
        file_bytes: Raw file content.
        file_format: One of json, jsonl, csv, xlsx.

    Returns:
        List of dictionaries, one per row/entry.
    """
    if file_format == "json":
        return _parse_json(file_bytes)
    if file_format == "jsonl":
        return _parse_jsonl(file_bytes)
    if file_format == "csv":
        return _parse_csv(file_bytes)
    if file_format == "xlsx":
        return _parse_xlsx(file_bytes)
    raise ValueError(f"Unsupported format: {file_format}")


def extract_headers_and_sample(
    file_bytes: bytes,
    file_format: str,
    max_rows: int = MAX_SAMPLE_ROWS,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Read just the headers and first N rows for the analyze step.

    For structured formats (JSON/JSONL) with nested objects, headers
    are derived from flattened top-level keys of the first entry.

    Returns:
        (headers, sample_rows)
    """
    rows = parse_file(file_bytes, file_format)
    if not rows:
        return [], []

    # Collect all unique keys across sample rows (preserving order)
    sample = rows[:max_rows]
    seen_keys: Dict[str, None] = {}
    for row in sample:
        for key in row:
            if key not in seen_keys:
                seen_keys[key] = None
    headers = list(seen_keys.keys())

    return headers, sample


# ── Internal parsers ─────────────────────────────────────────────


def _parse_json(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse JSON data – array, object with 'tests' key, or single object."""
    text = file_bytes.decode("utf-8")
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"Failed to parse JSON: {e}") from e

    if isinstance(data, list):
        return [entry for entry in data if isinstance(entry, dict)]

    if isinstance(data, dict):
        # Check for a top-level "tests" array
        if "tests" in data and isinstance(data["tests"], list):
            return [entry for entry in data["tests"] if isinstance(entry, dict)]
        # Single object
        return [data]

    raise ValueError(
        "JSON file must contain an array, an object with a 'tests' key, or a single object"
    )


def _parse_jsonl(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse newline-delimited JSON (one object per line)."""
    text = file_bytes.decode("utf-8")
    rows: List[Dict[str, Any]] = []
    for line_num, line in enumerate(text.splitlines(), 1):
        line = line.strip()
        if not line:
            continue
        try:
            entry = json.loads(line)
            if isinstance(entry, dict):
                rows.append(entry)
        except json.JSONDecodeError:
            logger.warning(f"Skipping invalid JSON on line {line_num}")
    return rows


def _parse_csv(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse CSV with header row.  Handles BOM encoding and auto-detects delimiter."""
    # Try utf-8-sig first (handles BOM), fall back to utf-8
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("utf-8")

    # Auto-detect delimiter (handles semicolons, tabs, pipes, etc.)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for row in reader:
        # DictReader can produce None keys for overflow fields; strip them out
        cleaned = {k: v for k, v in row.items() if k is not None and k != ""}
        # Skip entirely empty rows
        if not cleaned or all(v is None or str(v).strip() == "" for v in cleaned.values()):
            continue
        rows.append(cleaned)
    return rows


def _parse_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse Excel file using openpyxl, returning rows as dicts.

    Searches all sheets for the first one that contains data, and skips
    any leading blank rows before the header row.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel import. Install it with: pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Try the active sheet first, then fall back to the others
    sheets_to_try = []
    if wb.active is not None:
        sheets_to_try.append(wb.active)
    for name in wb.sheetnames:
        ws = wb[name]
        if ws not in sheets_to_try:
            sheets_to_try.append(ws)

    for ws in sheets_to_try:
        all_rows = list(ws.iter_rows(values_only=True))

        # Find the first non-empty row to use as headers
        header_idx = None
        for i, row in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header_idx = i
                break

        if header_idx is None:
            continue

        raw_headers = all_rows[header_idx]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        result = []
        for row in all_rows[header_idx + 1 :]:
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_dict = {}
            for header, value in zip(headers, row):
                if not header:
                    continue
                row_dict[header] = value
            if row_dict:
                result.append(row_dict)

        if result:
            wb.close()
            return result

    wb.close()
    return []
